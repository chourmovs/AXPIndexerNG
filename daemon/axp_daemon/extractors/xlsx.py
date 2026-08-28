from datetime import date, datetime, time

def _display(value):
    if isinstance(value, (datetime, date, time)):
        return value.isoformat(sep=" ") if isinstance(value, datetime) else value.isoformat()
    return str(value)


def extract(path):
    """Stream a workbook in read-only mode and retain sheet, row and cell coordinates."""
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:  # Hidden sheets intentionally remain searchable.
            lines = [f"Workbook: {path.name}", f"Sheet: {sheet.title}"]
            for row_number, row in enumerate(sheet.iter_rows(values_only=True), 1):
                cells = [f"{get_column_letter(column)}={_display(value)}"
                         for column, value in enumerate(row, 1) if value is not None and str(value) != ""]
                if cells:
                    lines.append(f"Row {row_number} | " + " | ".join(cells))
                if len(lines) >= 102:
                    yield "\n".join(lines), None
                    lines = [f"Workbook: {path.name}", f"Sheet: {sheet.title} (continued)"]
            if len(lines) > 2:
                yield "\n".join(lines), None
    finally:
        workbook.close()
